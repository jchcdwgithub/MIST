from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.table import Table
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
from typing import Dict, List, Tuple
import math

def write_tables_to_excel_workbook(tables_arrays:Dict[str,List[List[List[str]]]], output_file:str="AP Status.xlsx", workbook:str="") -> None:
    
    if workbook == "":
        workbook = Workbook()
    else:
        try:
            workbook = load_workbook(workbook)
        except FileNotFoundError as e:
            print(e)
            print('Check the file path/filename and try again.')
            exit()
    
    first_worksheet = True
    for worksheet_title in tables_arrays:
        if first_worksheet:
            current_worksheet = workbook.active
            current_worksheet.title = worksheet_title
            first_worksheet = False
        else:
            current_worksheet = workbook.create_worksheet(title=worksheet_title if len(worksheet_title) < 30 else worksheet_title[:30])
        current_row = 1
        for table in tables_arrays[worksheet_title]:
            current_row = add_table_to_worksheet(table, current_worksheet, start=current_row) 
        workbook.save(output_file)

def add_table_to_worksheet(data:List[List[str]],worksheet,table_name='',start:int=1) -> int:
    
    end_cell_number = len(data)
    end_cell = (end_cell_number, start + len(data[0]))
    start_cell = (1,start)
    num_columns = len(data[0])

    current = 1 
    data_index = 0
    while current <= end_cell_number:
        current_column = start
        current_row_cells = []
        while current_column < start + num_columns:
            current_row_cells.append(worksheet.cell(row=current, column=current_column))
            current_column += 1
        current_data_row = data[data_index]
        for cell,data_value in zip(current_row_cells,current_data_row):
            cell.value = data_value
        current += 1
        data_index += 1 
    add_color_scheme(worksheet,start_cell,end_cell)
    column_widths = get_widest_column_widths(start-1,data)
    adjust_column_widths(worksheet,column_widths)

def add_color_scheme(worksheet,start_cell:Tuple[int,int],end_cell:Tuple[int,int]) -> None:
    ''' Add a red table color scheme to the cells in the worksheet. '''
    title_row_color = 'C0504D'
    colors = ['E6B8B7', 'F2DCDB']
    color_index = 0
    rows = []
    start_row,start_column = start_cell
    end_row,end_column = end_cell
    current_row = start_row 
    while current_row <= end_row:
        current_column = start_column 
        current_row_list = []
        while current_column < end_column:
            current_row_list.append(worksheet.cell(row=current_row,column=current_column))
            current_column += 1
        rows.append(current_row_list)
        current_row += 1
    for row in rows:
       if color_index == 0:
          for cell in row:
               cell.fill = PatternFill('solid', fgColor = title_row_color)
               cell.font = Font(color='FFFFFF')
       else:
            for cell in row:
               cell.fill = PatternFill('solid', fgColor=colors[color_index%2])
               cell.font = Font(color='000000')
       color_index += 1

def get_widest_column_widths(start:int,data:List[str]) -> Dict[str,int]:
    ''' Returns a dictionary of column letter to widest length of data found in each column. '''
    column_widths = {}
    columns_table = swap_rows_and_columns(data)
    column_index = start
    for column in columns_table:
        column_letter = calculate_column_letters(column_index)
        current_longest_len = 0
        for cell in column:
            if len(cell) > current_longest_len:
                column_widths[column_letter] = len(cell)
                current_longest_len = len(cell)
        column_index += 1
    return column_widths

def adjust_column_widths(worksheet,column_widths:Dict[str,int]) -> None:
    ''' Adjust the column widths of the table in the worksheet based on the dictionary passed. '''

    for column_letter in column_widths:
        worksheet.column_dimensions[column_letter].width = column_widths[column_letter] + 5

def swap_rows_and_columns(data:List[List[str]]) -> List[List[str]]:
    ''' Given an array of arrays, swap the rows and columns data. '''

    num_rows = len(data)
    swapped_matrix = []
    num_cols = len(data[0])
    current_col = 0
    while current_col < num_cols:
        current_row = 0
        current_col_array = []
        while current_row < num_rows:
            current_col_array.append(data[current_row][current_col])
            current_row += 1
        current_col += 1
        swapped_matrix.append(current_col_array)

    return swapped_matrix

def calculate_column_letters(column_index:int) -> str:
    ''' Given a column index, return the corresponding excel column letter. '''

    column_letters = [letter.upper() for letter in 'abcdefghijklmnopqrstuvwxyz']

    if column_index < 26:
        return column_letters[column_index]
    else:
        first_letter_index = math.floor(column_index/26) - 1
        last_letter_index = column_index%26 - 1

        first_letter = column_letters[first_letter_index]
        last_letter = column_letters[last_letter_index]

        column = first_letter + last_letter
        return column